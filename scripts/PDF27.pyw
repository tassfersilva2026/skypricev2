# -*- coding: utf-8 -*-
# === PDF27.pyw (ADAPTADO P/ GITHUB) ===
# Alterações mínimas:
# - Paths relativos ao repo (inbox/, data/, out, state)
# - Execução de rodada única (--once) para Actions
# - Parquet por rodada: data/OFERTAS_YYYY-MM-DD_HH:MM:SS.parquet (hora SP)
# - XLS por rodada: out/OFERTAS_RUN_YYYY-MM-DD_HH-MM-SS.xlsx (artefato)
# - Cache simples de PDFs (mtime+tamanho) para não reextrair repetidos
# - Limite opcional de PDFs por rodada (--limit ou TEST_MAX_PDFS)

import os, re, glob, json, time, argparse, logging
from datetime import datetime
import pdfplumber
import pandas as pd
from tqdm import tqdm
from openpyxl.utils import get_column_letter
from zoneinfo import ZoneInfo

# ── CONFIGS ───────────────────────────────────────────────────────────────────
PDF_DIR        = os.environ.get("PDF_DIR",        "inbox")
MATRIX_XLSX    = os.environ.get("MATRIX_XLSX",    "out/OFERTASMATRIZ.xlsx")
PARQUET_OFS    = os.environ.get("PARQUET_OFS",    "out/OFERTASMATRIZ_OFERTAS.parquet")
PARQUET_ERR    = os.environ.get("PARQUET_ERR",    "out/OFERTASMATRIZ_ERROS.parquet")
SHEET_OFERTAS  = "OFERTAS"
SHEET_ERROS    = "ERRO_MONITORAMENTO"

ROW_IDS_FILE   = os.environ.get("ROW_IDS_FILE",   "state/OFERTASMATRIZ_ROW_IDS.txt")
ERR_IDS_FILE   = os.environ.get("ERR_IDS_FILE",   "state/OFERTASMATRIZ_ERR_IDS.txt")
STATE_JSON     = os.environ.get("STATE_JSON",     "state/OFERTASMATRIZ_STATE.json")
TZ_NAME        = os.environ.get("TZ_NAME",        "America/Sao_Paulo")

# Limite opcional para testar rodadas pequenas
ENV_MAX_PDFS   = os.environ.get("TEST_MAX_PDFS")
ENV_MAX_PDFS   = int(ENV_MAX_PDFS) if (ENV_MAX_PDFS or "").strip().isdigit() else None

# Parquet por rodada e XLS artefato
WRITE_SINGLE_RUN_PARQUET = os.environ.get("WRITE_SINGLE_RUN_PARQUET", "1").lower() in ("1","true","yes")
WRITE_XLS_ARTIFACT       = os.environ.get("WRITE_XLS_ARTIFACT", "1").lower() in ("1","true","yes")
RUN_TS                   = os.environ.get("RUN_TS")  # "YYYY-MM-DD_HH-MM-SS"

os.makedirs(os.path.dirname(MATRIX_XLSX), exist_ok=True)
os.makedirs(os.path.dirname(PARQUET_OFS), exist_ok=True)
os.makedirs(os.path.dirname(PARQUET_ERR), exist_ok=True)
os.makedirs("data", exist_ok=True)
os.makedirs("state", exist_ok=True)
os.makedirs("out", exist_ok=True)

logging.getLogger("pdfminer").setLevel(logging.ERROR)

VALID_ENTITIES = {
    "123milhas","agoda","airbnb","aircanada","airfrance","aeromexico","americanairlines",
    "ancoradouro","azul","booking.com","capoviagens","cestarollitravel","confianca",
    "cvc","decolar","esferatur","expedia","flipmilhas","flytourgapnet","gol",
    "googleflights","gotogate","hoteis.com","hurb","iberia","jetsmart","kayak",
    "kissandfly","kiwi.com","latam","lufthansa","maxmilhas","momondo","mrsmrssmith",
    "mytrip","passagenspromo","primetour","queensberryviagens","rexturadvance",
    "sakuratur","skyscanner","submarinoviagens","tap","trendoperadora","traveloka",
    "trip.com","unitedairlines","viajanet","visualturismo","voepass","vrbo","zarpo","zupper"
}
AIRLINES = ["gol","latam","azul","voepass","jetsmart","airfrance","unitedairlines",
            "iberia","lufthansa","aeromexico","aircanada","americanairlines","tap"]

# Somente TRECHOS permitidos
ALLOWED_TRECHOS = {
    "BELGRU","BSBGRU","CGHGIG","CGHREC","FORGRU","GIGCGH",
    "GRUMCZ","GRUREC","GRUSSA","RECGRU","SAOMCZ"
}

# Regras
dates_regex   = re.compile(r"(\d{2}/\d{2}/\d{4},\s*\d{2}:\d{2})")
price_regex   = re.compile(r"R\$\s*([\d\s\.,]+)")
CUTOFF_OFFERS = "complemente sua viagem"
TIMES_CUTOFF  = "verificando preços e disponibilidade"
FIRST_PAGE_ERROR_RULES = {
    re.compile(r"as melhores ofertas e promoções", re.IGNORECASE):  "ERRO DE PAGINA",
    re.compile(r"destinos nacionais mais buscados", re.IGNORECASE):  "ERRO DE PAGINA",
    re.compile(r"encaminhando para o website soli", re.IGNORECASE):  "ERRO DE PAGINA",
    re.compile(r"passagens aéreas em promoção\s*\|\s*l", re.IGNORECASE): "ERRO DE PAGINA",
    re.compile(r"skyscanner\s+você\s+é\s+uma\s+pessoa\s+ou", re.IGNORECASE): "ERRO ANTIBOT",
}
PAGE_ERROR_PATTERNS = {
    re.compile(r"passagens aéreas.*hotéis.*aluguel de carros", re.IGNORECASE): "ErroPaginaInicial",
    re.compile(r"pacotes de viagens", re.IGNORECASE): "ErroPaginaDecolar"
}
MONTH_MAP = {"jan":1,"fev":2,"mar":3,"abr":4,"mai":5,"jun":6,"jul":7,"ago":8,"set":9,"out":10,"nov":11,"dez":12}

# === Funções utilitárias (fieis ao espírito do seu script) ====================
def to_upper_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty: return df
    return df.applymap(lambda x: x.upper() if isinstance(x, str) else x)

def _filter_by_allowed_trechos(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty or "TRECHO" not in df.columns:
        return df
    return df[df["TRECHO"].astype(str).str.upper().isin(ALLOWED_TRECHOS)].copy()

def first_page_error_code(pdf_path):
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages: return None, None
            text = (pdf.pages[0].extract_text() or "")
    except Exception:
        return None, None
    low = text.lower()
    for pat, code in FIRST_PAGE_ERROR_RULES.items():
        m = pat.search(low)
        if m:
            start = max(0, m.start()-25); end = min(len(text), m.end()+25)
            trecho = text[start:end].replace("\n"," ").strip()
            return code, trecho
    return None, None

def extract_flight_info(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text() or ""
    low = text.lower()

    for pat, err in PAGE_ERROR_PATTERNS.items():
        if pat.search(text): return None, err

    idx = low.find(TIMES_CUTOFF)
    snippet = text[:idx] if idx != -1 else text

    all_times = re.findall(r"\b(\d{2}:\d{2})\b", snippet)
    times_dict = {f"Horário{i+1}": t for i, t in enumerate(all_times)}

    tipo = ""
    matches = list(re.finditer(r"\b(\d{2}:\d{2})\b", snippet))
    if len(matches) >= 2:
        pos2 = matches[1].end()
        window = snippet[pos2:pos2+200].lower()
        if re.search(r"\bdireto\b", window): tipo = "DIRETO"
        else:
            m_esc = re.search(r"(\d+)\s*escalas?", window)
            m_par = re.search(r"(\d+)\s*paradas?", window)
            if m_esc: tipo = f"{m_esc.group(1)} ESCALAS"
            elif m_par: tipo = f"{m_par.group(1)} PARADAS"

    cia = next((c.upper() for c in AIRLINES if c in low), "")

    dm = re.search(r"ida[^\d]*(\d{1,2})\s+de\s+([a-zç]+)\.?\s+de\s+(\d{4})", low)
    if dm:
        day, m_pt, yr = dm.groups()
        m = MONTH_MAP.get(m_pt[:3], 0)
        flight_date = f"{int(day):02d}/{m:02d}/{yr}"
    else:
        flight_date = ""

    return {"Companhia Aérea": cia, **times_dict, "Tipo de Voo": tipo, "Data do Voo": flight_date}, None

def extract_offers_from_pdf(pdf_path, search_dt):
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for p in pdf.pages:
            text += (p.extract_text() or "") + "\n"

    text = re.sub(r"(\d)\n(\d)", r"\1\2", text)
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    cutoff = next((i for i,l in enumerate(lines) if CUTOFF_OFFERS in l.lower()), len(lines))
    lines = lines[:cutoff]

    if not search_dt:
        for l in lines[:10]:
            m = dates_regex.search(l)
            if m: search_dt = m.group(1); break

    offers, last_ent = [], None
    for l in lines:
        norm = re.sub(r"\s+","", l.lower())
        for ent in VALID_ENTITIES:
            if ent.replace(" ","") in norm:
                last_ent = ent; break
        pm = price_regex.search(l)
        if pm and last_ent:
            raw = pm.group(1)
            num = re.sub(r"[^\d,]","", raw).replace(",", ".")
            try: price = float(num)
            except: continue
            offers.append({"Agência/Companhia": last_ent, "Preço": price})
            last_ent = None
    return offers, search_dt

def get_trecho(file_name):
    token = re.split(r"[_.-]", file_name)[0].upper()
    return token

def rank_prices(df):
    df['Ranking'] = df.groupby('Nome do Arquivo')['Preço'].rank(method='min', ascending=True).astype(int)
    return df

def todas_colunas_preenchidas(row, cols_req):
    return all(pd.notna(row.get(col)) and str(row.get(col)).strip() != "" for col in cols_req)

# === Nomes por rodada (Parquet + XLS) =========================================
def _run_names():
    """
    Retorna (parquet_path, xls_path) p/ a rodada atual.
    - Parquet: OFERTAS_YYYY-MM-DD_HH:MM:SS.parquet (hora SP)
    - XLS:     OFERTAS_RUN_YYYY-MM-DD_HH-MM-SS.xlsx (compatível Windows)
    """
    if RUN_TS:
        date_part, time_part = RUN_TS.split("_", 1)
        time_colon = time_part.replace("-", ":")
        parquet = f"data/OFERTAS_{date_part}_{time_colon}.parquet"
        xls     = f"out/OFERTAS_RUN_{RUN_TS}.xlsx"
    else:
        now_sp = datetime.now(ZoneInfo(TZ_NAME))
        parquet = f"data/OFERTAS_{now_sp.strftime('%Y-%m-%d')}_{now_sp.strftime('%H:%M:%S')}.parquet"
        xls     = f"out/OFERTAS_RUN_{now_sp.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    return parquet, xls

# === Estado para não reprocessar PDFs iguais (mtime + tamanho) ================
def load_state() -> dict:
    try:
        with open(STATE_JSON, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_state(state: dict):
    try:
        os.makedirs(os.path.dirname(STATE_JSON), exist_ok=True)
        with open(STATE_JSON, "w", encoding="utf-8") as f:
            json.dump(state, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

# ── Excel/Parquet ─────────────────────────────────────────────────────────────
def write_back_preserving(file_path, df_ofertas, df_erros):
    mode = "a" if os.path.exists(file_path) else "w"
    with pd.ExcelWriter(file_path, engine="openpyxl", mode=mode, if_sheet_exists=("replace" if mode=="a" else None)) as writer:
        (df_ofertas if df_ofertas is not None else pd.DataFrame()).to_excel(writer, index=False, sheet_name=SHEET_OFERTAS)
        (df_erros   if df_erros   is not None else pd.DataFrame()).to_excel(writer, index=False, sheet_name=SHEET_ERROS)
        try:
            ws = writer.sheets[SHEET_OFERTAS]
            cols = list(df_ofertas.columns) if df_ofertas is not None else []
            for col_name in ["Data do Voo","Data/Hora da Busca"]:
                if col_name in cols:
                    idx = cols.index(col_name) + 1
                    col_letter = get_column_letter(idx)
                    for cell in ws[col_letter][1:]:
                        cell.number_format = "DD/MM/YYYY"
        except Exception:
            pass

def export_xls_per_run(of_df: pd.DataFrame, er_df: pd.DataFrame | None):
    """Gera um XLS com abas OFERTAS e ERRO_MONITORAMENTO para artefato do run."""
    _, xls_path = _run_names()
    with pd.ExcelWriter(xls_path, engine="openpyxl", mode="w") as writer:
        # CORREÇÃO: sem avaliar DataFrame como booleano
        (of_df if of_df is not None else pd.DataFrame()).to_excel(writer, index=False, sheet_name=SHEET_OFERTAS)
        (er_df if er_df is not None else pd.DataFrame()).to_excel(writer, index=False, sheet_name=SHEET_ERROS)
        try:
            ws = writer.sheets[SHEET_OFERTAS]
            cols = list(of_df.columns) if of_df is not None else []
            for col_name in ["Data do Voo","Data/Hora da Busca"]:
                if col_name in cols:
                    idx = cols.index(col_name) + 1
                    col_letter = get_column_letter(idx)
                    for cell in ws[col_letter][1:]:
                        cell.number_format = "DD/MM/YYYY"
        except Exception:
            pass
    print(f"📄 XLS por rodada: {xls_path}")
    return xls_path

def export_parquet(ofertas_df, erros_df):
    of = ofertas_df.copy()
    of = _filter_by_allowed_trechos(of)

    for c in ["Data do Voo","Data/Hora da Busca"]:
        if c in of:
            of[c] = pd.to_datetime(of[c], errors="coerce")

    # Consolidado interno
    of.to_parquet(PARQUET_OFS, index=False)
    if erros_df is not None and not erros_df.empty:
        er = erros_df.copy()
        er.to_parquet(PARQUET_ERR, index=False)

    # Parquet por rodada no padrão solicitado (timezone SP)
    if WRITE_SINGLE_RUN_PARQUET:
        parquet_path, _ = _run_names()
        of.to_parquet(parquet_path, index=False)
        print(f"📦 Parquet por rodada: {parquet_path} (linhas={len(of)})")

    # XLS por rodada como artefato
    if WRITE_XLS_ARTIFACT:
        export_xls_per_run(of, erros_df)

# ── 1 ciclo de atualização ────────────────────────────────────────────────────
def run_cycle(max_files: int | None = None):
    pdfs = sorted(glob.glob(os.path.join(PDF_DIR, "*.pdf")))
    if not pdfs:
        print("Nenhum PDF na pasta. Saindo.")
        return pd.DataFrame(), pd.DataFrame(), 0, 0

    # estado mtime+tamanho
    state = load_state()
    def _sig(p: str) -> str | None:
        try:
            st = os.stat(p)
            return f"{st.st_size}-{int(st.st_mtime)}"
        except Exception:
            return None
    to_process = []
    for p in pdfs:
        sig = _sig(p)
        if sig is None: continue
        if state.get(p) != sig:
            to_process.append(p)

    if not to_process:
        print("Nada novo a processar (cache mtime+tamanho).")
        return pd.DataFrame(), pd.DataFrame(), 0, 0

    # aplica limite (teste)
    eff_limit = max_files if max_files is not None else ENV_MAX_PDFS
    if eff_limit is not None and eff_limit > 0:
        to_process = to_process[:eff_limit]
        print(f"[teste] Limitando processamento a {eff_limit} PDF(s).")

    offers_rows, errors_rows = [], []
    for path in tqdm(to_process, desc="Processando PDFs"):
        fn = os.path.basename(path)

        code, trecho = first_page_error_code(path)
        if code:
            errors_rows.append({"Nome do Arquivo": fn, "Erro": code, "Trecho": trecho, "Pagina": 1})

        flight_info, err = extract_flight_info(path)
        if err:
            errors_rows.append({"Nome do Arquivo": fn, "Erro": err, "Trecho": "", "Pagina": 1})

        offers, sdt = extract_offers_from_pdf(path, "")
        for o in offers:
            offers_rows.append({
                "Nome do Arquivo": fn,
                **(flight_info or {}),
                "Data/Hora da Busca": sdt,  # "09/08/2025, 13:13"
                **o,
                "TRECHO": get_trecho(fn)
            })

    new_offers_df = pd.DataFrame(offers_rows)
    new_erros_df  = pd.DataFrame(errors_rows)

    if not new_offers_df.empty:
        new_offers_df["Data do Voo"] = pd.to_datetime(new_offers_df["Data do Voo"], dayfirst=True, errors="coerce")
        so_data = new_offers_df["Data/Hora da Busca"].astype(str).str.extract(r"(\d{2}/\d{2}/\d{4})", expand=False)
        new_offers_df["Data/Hora da Busca"] = pd.to_datetime(so_data, dayfirst=True, errors="coerce")
        diff_days = (new_offers_df["Data do Voo"].dt.normalize() - new_offers_df["Data/Hora da Busca"].dt.normalize()).dt.days
        new_offers_df["ADVP"] = diff_days.fillna(0).astype(int)
        new_offers_df = rank_prices(new_offers_df)
        req = ["Nome do Arquivo","Companhia Aérea","Horário1","Horário2","Horário3",
               "Tipo de Voo","Data do Voo","Data/Hora da Busca",
               "Agência/Companhia","Preço","TRECHO","ADVP","Ranking"]
        new_offers_df = new_offers_df[new_offers_df.apply(lambda r: todas_colunas_preenchidas(r, req), axis=1)]
        new_offers_df = new_offers_df[new_offers_df["Agência/Companhia"].str.lower() != "skyscanner"]
        new_offers_df = to_upper_df(new_offers_df)

    if not new_erros_df.empty:
        new_erros_df = to_upper_df(new_erros_df)

    # Carrega base atual, se existir
    try: base_ofertas = pd.read_excel(MATRIX_XLSX, sheet_name=SHEET_OFERTAS, engine="openpyxl")
    except: base_ofertas = pd.DataFrame()
    try: base_erros = pd.read_excel(MATRIX_XLSX, sheet_name=SHEET_ERROS, engine="openpyxl")
    except: base_erros = pd.DataFrame()

    base_cols = list(base_ofertas.columns) if not base_ofertas.empty else \
        ["Nome do Arquivo","Companhia Aérea","Horário1","Horário2","Horário3",
         "Tipo de Voo","Data do Voo","Data/Hora da Busca","Agência/Companhia",
         "Preço","TRECHO","ADVP","Ranking"]
    err_cols = list(base_erros.columns) if not base_erros.empty else ["Nome do Arquivo","Erro","Trecho","Pagina","EM_AMBAS"]

    novos_unicos = new_offers_df.copy() if not new_offers_df.empty else pd.DataFrame()
    new_errs_unique = new_erros_df.copy() if not new_erros_df.empty else pd.DataFrame()

    if not novos_unicos.empty:
        for c in base_cols:
            if c not in novos_unicos.columns: novos_unicos[c] = pd.NA
        novos_unicos = novos_unicos[base_cols]
    if not new_errs_unique.empty:
        if "EM_AMBAS" not in new_errs_unique.columns: new_errs_unique["EM_AMBAS"] = pd.NA
        for c in err_cols:
            if c not in new_errs_unique.columns: new_errs_unique[c] = pd.NA
        new_errs_unique = new_errs_unique[err_cols]

    final_ofertas = pd.concat([base_ofertas, novos_unicos], ignore_index=True) if not novos_unicos.empty else base_ofertas
    final_erros   = pd.concat([base_erros, new_errs_unique], ignore_index=True) if not new_errs_unique.empty else base_erros

    if not final_erros.empty:
        ofertas_set = set(final_ofertas["Nome do Arquivo"].dropna().astype(str)) if "Nome do Arquivo" in final_ofertas else set()
        final_erros["EM_AMBAS"] = final_erros["Nome do Arquivo"].apply(lambda x: "SIM" if str(x) in ofertas_set else "NÃO")

    if not novos_unicos.empty or not new_errs_unique.empty or base_ofertas.empty or base_erros.empty:
        write_back_preserving(MATRIX_XLSX, final_ofertas, final_erros)
        export_parquet(final_ofertas, final_erros)

    # Atualiza cache mtime+tamanho
    state = load_state()
    for p in to_process:
        try:
            st = os.stat(p)
            state[p] = f"{st.st_size}-{int(st.st_mtime)}"
        except Exception:
            pass
    save_state(state)

    return final_ofertas, final_erros, len(novos_unicos), len(new_errs_unique)

# ── CLI / Execução ────────────────────────────────────────────────────────────
if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--once", action="store_true", help="Executa apenas 1 ciclo.")
    ap.add_argument("--limit", type=int, default=None, help="Limite de PDFs a processar na rodada.")
    args = ap.parse_args()

    max_files = args.limit if args.limit is not None else ENV_MAX_PDFS

    if args.once:
        of_df, er_df, n_of, n_er = run_cycle(max_files=max_files)
        print(f"✅ Ciclo concluído. Ofertas novas: {n_of} | Erros novos: {n_er}")
    else:
        while True:
            try:
                of_df, er_df, n_of, n_er = run_cycle(max_files=max_files)
                print(f"✅ Ciclo concluído. Ofertas novas: {n_of} | Erros novos: {n_er}")
            except Exception as e:
                print(f"❌ Erro no ciclo: {e}")
            time.sleep(600)
